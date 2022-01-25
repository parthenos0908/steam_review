import json
import csv
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from os import path

# ID = 227300
ID = 255710

ATTENTION_FILENAME = "data/" + str(ID) + "/" + str(ID) + "_result.xlsx"


def main():
    r_filename = path.join(path.dirname(
        __file__), "data/" + str(ID) + "/" + str(ID) + "_review_predict.json")
    with open(r_filename, mode='r') as f:
        review_pred = json.load(f)
    f_filename = path.join(path.dirname(
        __file__), "data/" + str(ID) + "/" + str(ID) + "_forum_predict.json")
    with open(f_filename, mode='r') as f:
        forum_pred = json.load(f)
    c_filename = path.join(path.dirname(
        __file__), "data/" + str(ID) + "/" + str(ID) + "_cross_predict.json")
    with open(c_filename, mode='r') as f:
        cross_pred = json.load(f)

    answer_list = []
    r_pred_list = []
    f_pred_list = []
    c_pred_list = []
    review_list = []
    for i in range(len(review_pred)):
        answer_list.append(review_pred[i]["answer"])
        r_pred_list.append(review_pred[i]["pred"])
        f_pred_list.append(forum_pred[i]["pred"])
        c_pred_list.append(cross_pred[i]["pred"])
        review_list.append(review_pred[i]["review"])

    attention_filename = path.join(path.dirname(__file__), ATTENTION_FILENAME)
    wb = openpyxl.Workbook()
    ws1 = wb.worksheets[0]

    ws1.column_dimensions['E'].width = 254.91

    ws1.cell(1, 1, value="answer")
    ws1.cell(1, 2, value="r_pread")
    ws1.cell(1, 3, value="f_pread")
    ws1.cell(1, 4, value="c_pread")
    ws1.cell(1, 5, value="review")
    for i in range(len(answer_list)):
        ws1.cell(i+2, 1, value=answer_list[i])
        ws1.cell(i+2, 2, value=r_pred_list[i])
        ws1.cell(i+2, 3, value=f_pred_list[i])
        ws1.cell(i+2, 4, value=c_pred_list[i])
        ws1.cell(i+2, 5, value=review_list[i])

    # 罫線
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)

    class_color = ["FFC0CB", "98FB98", "FFFACD"]  # ピンク, 薄緑, 黄色
    for row in ws1.iter_rows():
        for cell in row:
            if cell.value in [0, 1, 2]:
                cell.fill = openpyxl.styles.PatternFill(
                    patternType='solid', fgColor=class_color[cell.value], bgColor=class_color[cell.value])
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")  # 中央ぞろえ
                cell.font = openpyxl.styles.Font(
                    name="游ゴシック", bold=True, color="808080")
            else:
                cell.font = openpyxl.styles.Font(name="游ゴシック")
                cell.alignment = Alignment(wrapText=True)
            ws1[cell.coordinate].border = border

    ws1.auto_filter.ref = "A1:D{0}".format(len(review_pred) + 1)

    wb.save(attention_filename)


if __name__ == '__main__':
    main()
