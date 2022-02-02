from asyncore import write
from lib2to3.pgen2.token import AT
from pickle import FALSE, NONE
from tkinter import font
import numpy as np
import pandas as pd
import tensorflow as tf
import transformers
import json
import csv
import openpyxl
from os import path
import sys
import collections as cl
import random
from tensorflow.python.keras.utils.vis_utils import plot_model
import keras
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix, roc_curve, roc_auc_score, precision_recall_curve, auc
from sklearn.preprocessing import label_binarize
import numpy
from tqdm import tqdm
import matplotlib.pyplot as plt
import japanize_matplotlib  # matplotで日本語使える

# model_nameはここから取得(cf. https://huggingface.co/transformers/pretrained_models.html)
model_name = "bert-base-uncased"
tokenizer = transformers.BertTokenizer.from_pretrained(model_name)

# DLパラメータ
num_classes = 3
max_length = 128  # 256はメモリ不足
batch_size = 32
epochs = 30
hold_out_rate = 0.7  # 訓練データとテストデータの比率

is_learn = 1  # 新しく学習する(0) or 既存の学習結果使う(1)
is_del_less_words = False  # 1単語以下をテストデータから除外


def main():
    # review読み込み
    input_review_json_filename = path.join(
        path.dirname(__file__), INPUT_REVIEW_FILENAME)

    review_data = load_review_json(input_review_json_filename)
    random.seed(1)
    random.shuffle(review_data)

    # 訓練データ(review)
    train_data = review_data[:int(len(review_data)*hold_out_rate)]
    train_texts = []
    train_labels = []
    train_texts_origin = []
    for train_d in train_data:
        train_texts.append(train_d['review_lem'])
        train_labels.append(train_d['label'])
        train_texts_origin.append(train_d['review'])

    # テストデータ
    test_data = review_data[int(len(review_data)*hold_out_rate):]
    test_texts = []
    test_labels = []
    test_texts_origin = []
    # 単語数1以下を除去
    for test_d in test_data:
        if (test_d['num_words'] > 1) or (not is_del_less_words):
            test_texts.append(test_d['review_lem'])
            test_labels.append(test_d['label'])
            test_texts_origin.append(test_d['review'])

    # 教師データにforumを使う場合
    if MODE in ["forum", "cross"]:
        input_forum_json_filename = path.join(
            path.dirname(__file__), INPUT_FORUM_FILENAME)
        forum_data = load_forum_json(input_forum_json_filename)
        random.seed(1)
        random.shuffle(forum_data)
        train_texts = []
        train_labels = []
        train_texts_origin = []
        for forum_d in forum_data:
            train_texts.append(forum_d['combined'])
            train_labels.append(forum_d['label'])
            train_texts_origin.append(forum_d['title'])

    print("[train data] BR:{0}, FR:{1}, OTHER:{2} [test data] BR:{3}, FR:{4}, OTHER:{5}".format(
        train_labels.count(0), train_labels.count(1), train_labels.count(2), test_labels.count(0), test_labels.count(1), test_labels.count(2)))

# ===============================モデル構築===============================
    # x_:テキストデータ, y_:ラベル
    x_train = to_features(train_texts, max_length)
    y_train = tf.keras.utils.to_categorical(
        train_labels, num_classes=num_classes)
    x_valid = to_features(test_texts, max_length)
    y_valid = tf.keras.utils.to_categorical(
        test_labels, num_classes=num_classes)
    model = build_model(model_name, num_classes=num_classes,
                        max_length=max_length)

    output_model_filename = path.join(
        path.dirname(__file__), MODEL_WEIGHT_FILENAME)

    # 学習or既存のモデル使う
    if is_learn == 0:
        # 訓練
        es_cb = keras.callbacks.EarlyStopping(
            monitor='val_loss', patience=3, verbose=1, mode='auto', restore_best_weights=True)
        model.fit(x_train, y_train, batch_size=batch_size, epochs=epochs,
                  verbose=1, validation_data=(x_valid, y_valid), callbacks=[es_cb])

        # モデルの保存
        model.save_weights(output_model_filename)
    elif is_learn == 1:
        # モデルの読み込み
        model.load_weights(output_model_filename)


# ===============================予測(学習)===============================
    # x_test = to_features(test_texts, max_length)
    # y_test = np.asarray(test_labels)
    # y_test_b = label_binarize(y_test, classes=[0, 1, 2])
    # y_preda = model.predict(x_test)
    # score = y_preda
    # y_pred = np.argmax(score, axis=1)

# # ===============================予測(既存モデル利用)===============================
    x_test = to_features(test_texts, max_length)
    y_test = np.asarray(test_labels)
    y_test_b = label_binarize(y_test, classes=[0, 1, 2])
    y_preda = model.predict(x_test)
    # score[num_test_data][num_class(3)]
    score = y_preda[0]
    y_pred = np.argmax(score, axis=1)

# # ===============================attentionの可視化===============================

    # _attention[num_test_data][num_heads(12)][max_length][max_length]
    _attention = y_preda[1]
    attention = np.zeros((len(_attention), max_length))
    for i in range(len(_attention)):
        for j in range(12):
            attention[i] += _attention[i][j][0]
    ids, split_texts = to_ids(test_texts, max_length)

    attention_filename = path.join(path.dirname(__file__), ATTENTION_FILENAME)
    wb = openpyxl.Workbook()
    ws1 = wb.worksheets[0]
    ws1.title = "texts"
    ws2 = wb.create_sheet(title="attention")
    ws3 = wb.create_sheet(title="low text")

    for i in range(len(split_texts)):
        ws1.cell(2*i+1, 1, value=y_test[i])
        ws1.cell(2*i+1, 2, value=y_pred[i])
        ws2.cell(2*i+1, 1, value=y_test[i])
        ws2.cell(2*i+1, 2, value=y_pred[i])
        ws3.cell(2*i+1, 1, value=y_test[i])
        ws3.cell(2*i+1, 2, value=y_pred[i])
        ws3.cell(2*i+1, 3, value=test_texts_origin[i])
        ws3.cell(2*i+2, 3, value=test_texts[i])
        for j in range(len(split_texts[i])):
            ws1.cell(2*i+1, j+3, value=split_texts[i][j])
            ws2.cell(2*i+1, j+3, value=attention[i][j])

    for row in ws3.iter_rows():
        for cell in row:
            cell.font = openpyxl.styles.Font(name="游ゴシック")

    class_color = ["FFC0CB", "98FB98", "FFFACD"]  # ピンク, 薄緑, 黄色
    for row in ws1.iter_rows():
        for cell in row:
            if cell.value == "[PAD]":
                cell.font = openpyxl.styles.Font(
                    name="游ゴシック", color="d3d3d3", bold=True)  # d3d3d3:グレー
                ws2[cell.coordinate].font = openpyxl.styles.Font(
                    name="游ゴシック", color="d3d3d3", bold=True)
            else:
                cell.font = openpyxl.styles.Font(name="游ゴシック", bold=True)
                ws2[cell.coordinate].font = openpyxl.styles.Font(
                    name="游ゴシック", bold=True)
            if cell.column < 3 and cell.value in [0, 1, 2]:
                cell.fill = openpyxl.styles.PatternFill(
                    patternType='solid', fgColor=class_color[cell.value], bgColor=class_color[cell.value])
                ws2[cell.coordinate].fill = openpyxl.styles.PatternFill(
                    patternType='solid', fgColor=class_color[cell.value], bgColor=class_color[cell.value])
                ws3[cell.coordinate].fill = openpyxl.styles.PatternFill(
                    patternType='solid', fgColor=class_color[cell.value], bgColor=class_color[cell.value])
            if cell.column > 3 and not (cell.value in ["[CLS]", "[SEP]", "[PAD]"]) and cell.value:
                bg_color = to_hex_rgb(255, int(255*(1.0-min(1.0, ws2[cell.coordinate].value))), int(
                    255*(1.0-min(1.0, ws2[cell.coordinate].value))))
                cell.fill = openpyxl.styles.PatternFill(
                    patternType='solid', fgColor=bg_color, bgColor=bg_color)
                ws2[cell.coordinate].fill = openpyxl.styles.PatternFill(
                    patternType='solid', fgColor=bg_color, bgColor=bg_color)

    wb.save(attention_filename)

# ===============================AUC plot===============================
    true_list = [[], [], []]
    for label in y_test_b:
        true_list[0].append(label[0])
        true_list[1].append(label[1])
        true_list[2].append(label[2])

    score_list = [[], [], []]
    for s in score:
        score_list[0].append(s[0])
        score_list[1].append(s[1])
        score_list[2].append(s[2])

    auc = plot_roc(true_list, score_list)
    plot_pr(true_list, score_list)

# ===============================log===============================

    print("Accuracy: %.5f" % accuracy_score(y_test, y_pred))
    print(classification_report(y_test, y_pred))
    print(confusion_matrix(y_test, y_pred, labels=[0, 1, 2]))

    log_filename = path.join(
        path.dirname(__file__), LOG_FILENAME)
    with open(log_filename, "w") as f:
        print("[train data] BR:{0}, FR:{1}, OTHER:{2} [test data] BR:{3}, FR:{4}, OTHER:{5}".format(
            train_labels.count(0), train_labels.count(1), train_labels.count(2), test_labels.count(0), test_labels.count(1), test_labels.count(2)), file=f)
        print("Accuracy: %.5f" % accuracy_score(y_test, y_pred), file=f)
        print(classification_report(y_test, y_pred), file=f)

        print("AUC_BR:{0:.2f}".format(auc[0]), file=f)
        print("AUC_FR:{0:.2f}".format(auc[1]), file=f)
        print("AUC_OTHER:{0:.2f}".format(auc[2]), file=f)
        print("", file=f)

        print(confusion_matrix(y_test, y_pred, labels=[0, 1, 2]), file=f)

        print("---------------------------------", file=f)
        print("max_length:{0}".format(max_length), file=f)
        print("batch_size:{0}".format(batch_size), file=f)
        print("hold_out_rate:{0}".format(hold_out_rate), file=f)
        print("is_learn:{0}".format(is_learn), file=f)
        print("is_del_less_words:{0}".format(is_del_less_words), file=f)

    output_json_filename = path.join(path.dirname(__file__), OUTPUT_FILENAME)
    output_json(output_json_filename, test_texts,
                y_test, y_pred, score_list, test_texts_origin)


def to_hex_rgb(r, g, b):
    return str.format('{:02x}{:02X}{:02X}', r, g, b)


def plot_roc(true_list, score_list, TAGS=["バグ報告", "機能要求", "その他　"], COLOR=["#d62728", "#2ca02c", "#1f77b4"], fontsize=16):
    auc = []
    plt.rcParams["font.size"] = fontsize
    for i in range(len(TAGS)):
        fpr, tpr, thresholds = roc_curve(true_list[i], score_list[i])
        _auc = roc_auc_score(true_list[i], score_list[i])
        label = "{0}：AUC = {1:.2f}".format(
            TAGS[i], _auc)
        auc.append(_auc)
        plt.plot(fpr, tpr, label=label, color=COLOR[i])
        plt.xlabel('FPR: False positive rate')
        plt.ylabel('TPR: True positive rate')
        plt.ylim([0, 1])
        plt.grid()
    plt.plot([0, 1], [0, 1], 'k--', lw=1)
    plt.legend(loc="lower right")
    plt.tight_layout()
    roc_filename = path.join(path.dirname(__file__), ROC_FILENAME + ".pdf")
    plt.savefig(roc_filename)
    plt.clf()  # 描画リセット

    return auc


def plot_pr(true_list, score_list, TAGS=["バグ報告", "機能要求", "その他　"], COLOR=["#d62728", "#2ca02c", "#1f77b4"], fontsize=16):
    plt.rcParams["font.size"] = fontsize
    plt.tight_layout()
    for i in range(len(TAGS)-1):
        precision, recall, thresholds = precision_recall_curve(
            true_list[i], score_list[i])
        label = "[{0}] AUC = {1:.2f}".format(TAGS[i], auc(recall, precision))
        plt.plot(recall, precision, label=label, color=COLOR[i])
        plt.xlabel('recall')
        plt.ylabel('precision')
        plt.ylim([0, 1])
        plt.grid()
    plt.plot([1, 0], [1, 1], 'k--', lw=1)
    plt.legend(loc="lower left")
    plt.tight_layout()
    pr_filename = path.join(path.dirname(__file__), PR_FILENAME + ".pdf")
    plt.savefig(pr_filename)


# テキストのリストをtransformers用の入力データに変換


def to_features(texts, max_length):
    shape = (len(texts), max_length)
    # input_idsやattention_mask, token_type_idsの説明はglossaryに記載(cf. https://huggingface.co/transformers/glossary.html)
    input_ids = np.zeros(shape, dtype="int32")
    attention_mask = np.zeros(shape, dtype="int32")
    token_type_ids = np.zeros(shape, dtype="int32")
    for i, text in enumerate(tqdm(texts)):
        encoded_dict = tokenizer.encode_plus(
            text, max_length=max_length, pad_to_max_length=True)
        input_ids[i] = encoded_dict["input_ids"]
        attention_mask[i] = encoded_dict["attention_mask"]
        token_type_ids[i] = encoded_dict["token_type_ids"]
    return [input_ids, attention_mask, token_type_ids]

# attention出力用にto_featuresを書き換え


def to_ids(texts, max_length):
    shape = (len(texts), max_length)
    # input_idsやattention_mask, token_type_idsの説明はglossaryに記載(cf. https://huggingface.co/transformers/glossary.html)
    input_ids = np.zeros(shape, dtype="int32")
    split_text = [[] for i in range(len(texts))]
    for i, text in enumerate(tqdm(texts)):
        encoded_dict = tokenizer.encode_plus(
            text, max_length=max_length, pad_to_max_length=True)
        input_ids[i] = encoded_dict["input_ids"]
        split_text[i] = tokenizer.convert_ids_to_tokens(input_ids[i])
    return input_ids, split_text

# 単一テキストをクラス分類するモデルの構築


METRICS = [
    "acc",
    keras.metrics.Precision(name='precision'),
    keras.metrics.Recall(name='recall'),
    keras.metrics.AUC(name='auc'),
]


# def build_model(model_name, num_classes, max_length):
#     input_shape = (max_length, )
#     input_ids = tf.keras.layers.Input(input_shape, dtype=tf.int32)
#     attention_mask = tf.keras.layers.Input(input_shape, dtype=tf.int32)
#     token_type_ids = tf.keras.layers.Input(input_shape, dtype=tf.int32)
#     bert_model = transformers.TFBertModel.from_pretrained(
#         model_name, output_attentions=True)
#     last_hidden_state, pooler_output, attention_output = bert_model.bert(
#         input_ids,
#         attention_mask=attention_mask,
#         token_type_ids=token_type_ids
#     )
#     # transformersのバージョンによってはエラーが起きる．ver2.11.0で実行
#     score = tf.keras.layers.Dense(
#         num_classes, activation="softmax")(pooler_output)
#     model = tf.keras.Model(
#         inputs=[input_ids, attention_mask, token_type_ids], outputs=[score])
#     optimizer = tf.keras.optimizers.Adam(
#         learning_rate=3e-5, epsilon=1e-08, clipnorm=1.0)
#     model.compile(optimizer=optimizer,
#                   loss="categorical_crossentropy", metrics=METRICS)
#     return model

def build_model(model_name, num_classes, max_length):
    input_shape = (max_length, )
    input_ids = tf.keras.layers.Input(input_shape, dtype=tf.int32)
    attention_mask = tf.keras.layers.Input(input_shape, dtype=tf.int32)
    token_type_ids = tf.keras.layers.Input(input_shape, dtype=tf.int32)
    bert_model = transformers.TFBertModel.from_pretrained(
        model_name, output_attentions=True)
    last_hidden_state, pooler_output, attention_output = bert_model.bert(
        input_ids,
        attention_mask=attention_mask,
        token_type_ids=token_type_ids
    )
    # transformersのバージョンによってはエラーが起きる．ver2.11.0で実行
    score = tf.keras.layers.Dense(
        num_classes, activation="softmax")(pooler_output)
    model = tf.keras.Model(
        inputs=[input_ids, attention_mask, token_type_ids], outputs=[score, attention_output[-1]])
    optimizer = tf.keras.optimizers.Adam(
        learning_rate=3e-5, epsilon=1e-08, clipnorm=1.0)
    model.compile(optimizer=optimizer,
                  loss="categorical_crossentropy", metrics=METRICS)
    return model

# jsonを読み込んでリストに格納


def load_review_json(json_filename):
    review_labeled = []
    with open(json_filename, mode='r') as f:
        json_data = json.load(f)
    for text in json_data:
        if ("label" in text) and (text["label"] in [0, 1, 2]):
            review_labeled.append(text)
    return review_labeled


def load_forum_json(json_filename):
    forum_list = []
    with open(json_filename, mode='r') as f:
        json_data = json.load(f)
    for text in json_data:
        if text["num_words"] <= max_length:
            forum_list.append(text)
    return forum_list

# 予測結果をjsonに書き込み


def output_json(json_filename, comment_list, answer_list, pred_list, score_list, origin_list):
    output = []
    for i in range(len(comment_list)):
        data = cl.OrderedDict()
        data["review_lem"] = str(comment_list[i])
        data["review"] = str(origin_list[i])
        data["answer"] = int(answer_list[i])
        data["pred"] = int(pred_list[i])
        data["score"] = (float(score_list[0][i]), float(
            score_list[1][i]), float(score_list[2][i]))
        output.append(data)

    with open(json_filename, mode='w') as f:
        json.dump(output, f, sort_keys=True, indent=4)


# classifier.py [ID1] [ID2] [MODE]
# MODE:"r" or "f" or "c"
# ID: 227300(Eur truck sim), 255710(cities:Skylines)
if __name__ == '__main__':
    args = sys.argv
    if 4 <= len(args):
        if args[1].isdigit():
            ID1 = args[1]
            ID2 = args[2]
            if args[3] in ["r", "f", "c"]:
                if args[3] == "r":
                    MODE = "review"
                if args[3] == "f":
                    MODE = "forum"
                if args[3] == "c":
                    MODE = "cross"

                INPUT_REVIEW_FILENAME = "data/" + \
                    str(ID1) + "/" + str(ID1) + "_review_cleaned_out.json"
                INPUT_FORUM_FILENAME = "data/" + \
                    str(ID2) + "/" + str(ID2) + "_forum_cleaned.json"

                OUTPUT_FILENAME = "data/" + \
                    str(ID1) + "/" + str(ID1) + "_" + \
                    str(MODE) + "_predict" + ".json"
                MODEL_WEIGHT_FILENAME = "data/" + \
                    str(ID1) + "/" + str(ID1) + "_" + \
                    str(MODE) + "_model" + "/checkpoint"
                ATTENTION_FILENAME = "data/" + \
                    str(ID1) + "/" + str(ID1) + "_" + \
                    str(MODE) + "_attention.xlsx"
                ROC_FILENAME = "data/" + \
                    str(ID1) + "/" + str(ID1) + "_" + str(MODE) + "_ROC"
                PR_FILENAME = "data/" + \
                    str(ID1) + "/" + str(ID1) + "_" + str(MODE) + "_PR"
                LOG_FILENAME = "data/" + \
                    str(ID1) + "/" + str(ID1) + "_" + str(MODE) + "_log.txt"
                main()
            else:
                print('Argument must be "f" or "r" or "c"')
        else:
            print('Argument is not digit')
    else:
        print('Arguments are too short')
